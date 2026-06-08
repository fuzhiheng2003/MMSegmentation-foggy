import os
import torch
import sys
import inspect
from thop import profile
# from memory_profiler import profile as mprofile
from memory_profiler import memory_usage
from fusion_net import fusion_net
import time
import openpyxl
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side # 导入填充模块
# from torchvision import models
# from gpu_mem_track import MemTracker  # 引用显存跟踪代码
import tracemalloc

# @mprofile
def pad2affine(image, mod):
    if image.shape[-2] % mod!=0:
        padnum=int(((int(image.shape[-2]/mod)+1)*mod-image.shape[-2])*0.5)
        pad=torch.nn.ReflectionPad2d((0,0,padnum,padnum))
        image=pad(image)
    if image.shape[-1] % mod!=0:
        padnum=int(((int(image.shape[-1]/mod)+1)*mod-image.shape[-1])*0.5)
        pad=torch.nn.ReflectionPad2d((padnum,padnum,0,0))
        image=pad(image)
    return image
def check_and_create_excel(filename):
    # 检查文件是否存在
    if not os.path.exists(filename):
        print(f"文件 {filename} 不存在。")
        # 创建一个新的工作簿
        wb = Workbook()
        ws = wb.active
        # 中文表头
        ws.cell(row=1, column=1, value='模型')
        ws.cell(row=1, column=2, value='输入图像尺寸')
        ws.cell(row=1, column=3, value='输出图像尺寸')
        ws.cell(row=1, column=4, value='图片宽')
        ws.cell(row=1, column=5, value='图片高')
        ws.cell(row=1, column=6, value='缩放倍率')
        ws.cell(row=1, column=7, value='乘加累积数')
        ws.cell(row=1, column=8, value='计算量')
        ws.cell(row=1, column=9, value='参数量')
        ws.cell(row=1, column=10, value='运行时间（单位秒）')
        ws.cell(row=1, column=11, value='平均占用内存')
        ws.cell(row=1, column=12, value='峰值内存消耗')
        # 英文表头
        ws.cell(row=2, column=1, value='model')
        ws.cell(row=2, column=2, value='input image size')
        ws.cell(row=2, column=3, value='output image size')
        ws.cell(row=2, column=4, value='width')
        ws.cell(row=2, column=5, value='height')
        ws.cell(row=2, column=6, value='scale')
        ws.cell(row=2, column=7, value='macs(G)')
        ws.cell(row=2, column=8, value='flops(G)')
        ws.cell(row=2, column=9, value='params(M)')
        ws.cell(row=2, column=10, value='time (s)')
        ws.cell(row=2, column=11, value='average memory(MB)')
        ws.cell(row=2, column=12, value='max memory(MB)')
        ws.cell(row=3, column=1, value='IDB')
        # 保存到指定的文件名
        wb.save(filename)

        wb = load_workbook(filename)
        ws = wb.active
        # 获取最后一行的行号
        max_row = ws.max_row-1
        # 生成输入输出图像尺寸数据
        temp = 0
        data = [1280, 1024, 960, 768, 640, 512, 384, 256, 192, 128, 64, 32, 16]
        for col_num, cell_value in enumerate(data, 1):
            ws.cell(row=max_row + 1, column=2).value = f"{cell_value}x{cell_value}"
            ws.cell(row=max_row + 2, column=2).value = f"{int(cell_value/2)}x{int(cell_value/2)}"
            ws.cell(row=max_row + 3, column=2).value = f"{int(cell_value/4)}x{int(cell_value/4)}"
            ws.cell(row=max_row + 1, column=3).value = f"{cell_value}x{cell_value}"
            ws.cell(row=max_row + 2, column=3).value = f"{cell_value}x{cell_value}"
            ws.cell(row=max_row + 3, column=3).value = f"{cell_value}x{cell_value}"

            # 隔3行上色
            # 设置填充颜色
            fille = PatternFill('solid', fgColor='B9D3EE')
            # 定义灰色作为边框颜色
            light_gray = Side(style='thin')
            thin_border = Border(left=light_gray,
                                 right=light_gray,
                                 top=light_gray,
                                 bottom=light_gray)
            # 隔3行上色
            if temp%2 == 0:
                for j in range(1, 12 + 1):  # 遍历当前行的所有表格
                    ws.cell(row=max_row+1, column=j).fill = fille  # 将当前行的每一个表格填充颜色
                    ws.cell(row=max_row + 2, column=j).fill = fille  # 将当前行的每一个表格填充颜色
                    ws.cell(row=max_row + 3, column=j).fill = fille  # 将当前行的每一个表格填充颜色
                    ws.cell(row=max_row + 1, column=j).border = thin_border  # 设置当前格的边框格式
                    ws.cell(row=max_row + 2, column=j).border = thin_border  # 设置当前格的边框格式
                    ws.cell(row=max_row + 3, column=j).border = thin_border  # 设置当前格的边框格式
            else:
                for j in range(1, 12 + 1):  # 遍历当前行的所有表格
                    ws.cell(row=max_row + 1, column=j).border = thin_border  # 设置当前格的边框格式
                    ws.cell(row=max_row + 2, column=j).border = thin_border  # 设置当前格的边框格式
                    ws.cell(row=max_row + 3, column=j).border = thin_border  # 设置当前格的边框格式
            max_row += 3
            temp += 1
        # 存储文件
        wb.save(fileName)

        # 表格列名，目前是12列，有需求可以再加
        letter = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L']
        set_column_width(fileName, letter[1], 20)
        set_column_width(fileName, letter[2], 20)
        # 修改表格宽度
        for i in range(6):
            set_column_width(fileName, letter[6 + i], 20)
        print(f"已创建新的文件 {filename}")
    else:
        print(f"文件 {filename} 已存在。")

def set_column_width(filename, column_letter, width):
    # 加载工作簿
    wb = load_workbook(filename)
    # 假定我们在处理第一个工作表
    ws = wb.active
    # 设置指定列的宽度
    ws.column_dimensions[column_letter].width = width
    # 保存更改后的文件
    wb.save(filename)
if __name__ == '__main__':
    # scales = [1, 2, 4]
    # for i in scales:
    # 接收命令行参数
    width = int(sys.argv[1])
    height = int(sys.argv[2])

    # scale = i #倍数
    scale = int(sys.argv[3])

    device="cpu"

    input = torch.randn(1, 3, int(width/scale), int(height/scale)).to(device)
    input = pad2affine(input, mod=32)
    model = fusion_net(scale=scale,device=device).to(device)

    interval = 0.1  # 采样间隔时间，单位秒
    mem_samples = []

    start = time.time() # 记录开始时间
    flops, params = profile(model, inputs=(input,))
    end = time.time() # 记录结束时间
    # 以0.1秒一次的频率获取当前内存使用量,将其放入数组当中
    mem_usage = memory_usage((profile, (model, (input,),)), max_usage=False, interval=interval)
    #print(mem_usage)

    print(f"macs = {flops*2/1e9}G")
    print(f"params = {params/1e6}M")
    print('flops: ', flops/1e9, 'params: ', params/1e6)
    print("time_cost_thop:{}s".format((end-start)))
    # 计算平均内存使用量
    average_memory = sum(mem_usage) / len(mem_usage)
    # 保留五位小数输出
    print(f"average_memory:{round(average_memory,5)}MB = {round(average_memory/1024,5)}GB")
    print(f"max_memory:{round(max(mem_usage),5)}MB = {round(max(mem_usage)/1024,5)}GB")

    data = []
    data.append(width)
    data.append(height)
    data.append(scale)
    data.append(flops * 2 / 1e9)# macs
    data.append(flops / 1e9)# flops
    data.append(params/1e6)# params

    data.append(end-start)# time_cost_thop
    data.append(round(average_memory,5))# average_memory
    data.append(round(max(mem_usage),5))# max_memory
    # 文件名
    fileName = 'IDB.xlsx'
    check_and_create_excel(fileName)

    wb = load_workbook('./IDB.xlsx')

    ws = wb.active
    # 获取命令行指定行号
    max_row = int(sys.argv[4])
    # 每次运行都会在文件中追加写入测试数据
    for col_num, cell_value in enumerate(data, 1):
        ws.cell(row=max_row, column=col_num+3).value = cell_value
    wb.save(fileName)
    print(f"图片尺寸:{height}x{width} scale:{scale} 测试数据已写入文件IDB.xlsx")